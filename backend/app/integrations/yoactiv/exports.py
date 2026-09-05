"""Yoactiv **export-file** ingestion — a temporary bridge, not a replacement.

The Data API connector (``client.py`` / ``sync.py``) remains the intended
production path. It is blocked on credentials (see ``docs/INTEGRATIONS.md``),
so this module reads the same operational facts out of the two Excel/CSV
reports SLAM can export from the Yoactiv web console today:

======================  ===================================================
Yoactiv report          What it gives GymFlow
======================  ===================================================
Membership Report       members + their membership terms (one row per
(``memshiprpt.aspx``)   billed service): Member ID, Member Name, Mobile,
                        Service Name, Bill No, Start/End Date, Bill Amount,
                        Pay Mode, Lead Source, Sales Rep, Attendance Id
Member Check-ins        attendance: Member ID, Name, Mobile, Date, Clock
(``clientcheckins.aspx``)  In/Out, Location, Service Name, Medium/Staff
======================  ===================================================

**Everything downstream of parsing is shared with the API connector**, on
purpose — an export-imported member, membership or check-in must be
indistinguishable from an API-synced one:

* :func:`app.integrations.yoactiv.mapping.parse_dmy` / ``parse_clock`` parse
  the same ``dd-MM-yyyy`` / ``hh:mm AM/PM`` formats.
* :func:`app.integrations.yoactiv.identity.resolve_member` decides which
  GymFlow member a row belongs to, with the same precedence
  (``external_ref`` → exact email → unique active phone → ambiguous → none)
  and the same refusal to ever match on a name.
* :func:`app.integrations.yoactiv.lifecycle.apply_invoice` turns a billed
  service into ``Membership`` rows, so a renewal is a new row and
  ``Member.is_active`` is recomputed identically.

One genuine incompatibility the export surfaced, handled explicitly rather
than papered over: the check-in export carries **no ``Service_card_id``**,
which is part of the API's composed ``external_event_id``. An export row and
the API row for the very same visit therefore hash differently. Writing both
would double-count attendance. :func:`attendance_natural_key_exists` is the
shared guard against that — a check on
``(user_id, work_date, event_type, occurred_at)`` that both this module and
``sync.py`` consult before inserting. See ``docs/INTEGRATIONS.md``.

Nothing here invents a member, a membership or a visit. A row that cannot be
resolved is classified and reported, never guessed at.
"""

from __future__ import annotations

import csv
import hashlib
import io
import re
from collections import defaultdict
from collections.abc import Callable
from dataclasses import dataclass, field
from datetime import date, datetime, time
from enum import Enum
from html.parser import HTMLParser
from pathlib import Path
from typing import Any

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.v1.auth import normalise_phone
from app.core.clock import combine_branch
from app.db.models import (
    AttendanceEvent,
    Branch,
    CaptureMethod,
    EventType,
    Member,
    PersonType,
    User,
)
from app.integrations.yoactiv import identity, lifecycle
from app.integrations.yoactiv.mapping import (
    YoactivBilledService,
    YoactivInvoice,
    parse_clock,
    parse_dmy,
)

#: Marks every row this bridge writes, so an operator (and a later API sync)
#: can tell export-sourced data from API-sourced data.
SOURCE = "yoactiv-export"


class ExportKind(str, Enum):
    MEMBERSHIP = "membership"
    CHECKINS = "checkins"


class HeaderValidationError(ValueError):
    """The file's header row is not one of the two supported Yoactiv reports.

    Raised instead of proceeding on a partial or guessed match — a silently
    misparsed export is worse than a refusal.
    """


# --------------------------------------------------------------- header mapping
#
# Canonical field -> accepted header spellings, lower-cased and whitespace
# collapsed before comparison. Every spelling here was read off the real
# Yoactiv reports; none is invented. Extra columns in the file are ignored,
# so Yoactiv adding one does not break the import.

_MEMBERSHIP_COLUMNS: dict[str, tuple[str, ...]] = {
    "member_id": ("member id", "memberid", "member_id"),
    "member_name": ("member name", "name", "client name"),
    "mobile": ("mobile", "mobile number", "phone"),
    "service_name": ("service name", "service", "membership"),
    "start_date": ("start date", "startdate", "from date"),
    "end_date": ("end date", "enddate", "expiry date", "to date"),
}
_MEMBERSHIP_OPTIONAL: dict[str, tuple[str, ...]] = {
    "attendance_id": ("attendance id", "attendanceid"),
    "bill_no": ("bill no", "billno", "bill number", "invoice no"),
    "bill_amount": ("bill amount", "amount", "final amount"),
    "pay_mode": ("pay mode", "paymode", "payment mode"),
    "lead_source": ("lead source", "leadsource", "source"),
    "sales_rep": ("sales rep name", "sales rep", "salesrep"),
    "last_checkin": ("last check-in date", "last checkin date", "last check in date"),
    "email": ("email", "mail id", "e-mail", "email id"),
    "branch": ("branch", "location", "club", "centre", "center"),
}

_CHECKIN_COLUMNS: dict[str, tuple[str, ...]] = {
    "member_id": ("member id", "memberid", "member_id"),
    "member_name": ("name", "member name", "client name"),
    "date": ("date", "attendance date", "check-in date", "checkin date"),
}
_CHECKIN_OPTIONAL: dict[str, tuple[str, ...]] = {
    "mobile": ("mobile", "mobile number", "phone"),
    "clock_in": ("clock in", "clockin", "check-in", "in time"),
    "clock_out": ("clock out", "clockout", "check-out", "out time"),
    "location": ("location", "branch", "club", "centre", "center"),
    "service_name": ("service name", "service"),
    "medium": ("medium/staff", "medium", "staff", "medium / staff"),
    "conducted_by": ("conducted by", "conductedby", "pt staff"),
    "email": ("email", "mail id", "e-mail", "email id"),
}


def _norm_header(value: Any) -> str:
    """Compare header text case-, space- and underscore-insensitively.

    The on-screen report and its Excel export do not agree on spelling: the
    table shows "Member ID" and "Service Name", the workbook writes "MemberID"
    and "Service_Name". Underscores become spaces so one alias covers both.
    """
    return " ".join(str(value or "").replace("_", " ").split()).strip().lower()


def _index_headers(header_row: list[Any]) -> dict[str, int]:
    by_text: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        text = _norm_header(cell)
        if text:
            by_text.setdefault(text, idx)  # first occurrence wins
    return by_text


def _match(by_text: dict[str, int], aliases: tuple[str, ...]) -> int | None:
    for alias in aliases:
        if alias in by_text:
            return by_text[alias]
    return None


def detect_kind(header_row: list[Any]) -> ExportKind:
    """Which of the two supported reports this file is.

    Decided on the columns that actually distinguish them: a membership export
    carries a Start *and* an End Date; a check-in export carries a single
    Date. Clock columns are **not** part of the test — the on-screen check-in
    report shows Clock In / Clock Out but the Excel export omits them, so
    requiring one would reject the real file. Never guessed from the filename.
    """
    by_text = _index_headers(header_row)
    has_member = _match(by_text, _MEMBERSHIP_COLUMNS["member_id"]) is not None
    has_terms = (
        _match(by_text, _MEMBERSHIP_COLUMNS["start_date"]) is not None
        and _match(by_text, _MEMBERSHIP_COLUMNS["end_date"]) is not None
    )
    # "Last Check-In Date" on the membership report normalises to
    # "last check-in date", so it never collides with this.
    has_visit = _match(by_text, _CHECKIN_COLUMNS["date"]) is not None
    if has_member and has_terms:
        return ExportKind.MEMBERSHIP
    if has_member and has_visit:
        return ExportKind.CHECKINS
    raise HeaderValidationError(
        "This does not look like a Yoactiv Membership Report or Member Check-ins "
        "export. Expected either 'Member ID' + 'Start Date' + 'End Date', or "
        "'Member ID' + 'Date'. Found: "
        + ", ".join(sorted(t for t in by_text if t)[:15])
        + ". Refusing to guess — see docs/INTEGRATIONS.md."
    )


def resolve_headers(header_row: list[Any], kind: ExportKind) -> dict[str, int]:
    required = _MEMBERSHIP_COLUMNS if kind is ExportKind.MEMBERSHIP else _CHECKIN_COLUMNS
    optional = _MEMBERSHIP_OPTIONAL if kind is ExportKind.MEMBERSHIP else _CHECKIN_OPTIONAL
    by_text = _index_headers(header_row)

    resolved: dict[str, int] = {}
    missing: list[str] = []
    for canonical, aliases in required.items():
        idx = _match(by_text, aliases)
        if idx is None:
            missing.append(aliases[0])
        else:
            resolved[canonical] = idx
    if missing:
        raise HeaderValidationError(
            f"Yoactiv {kind.value} export is missing required column(s): "
            + ", ".join(missing)
            + ". Refusing to guess — check the export's header row."
        )
    for canonical, aliases in optional.items():
        idx = _match(by_text, aliases)
        if idx is not None:
            resolved[canonical] = idx
    return resolved


# ------------------------------------------------------------------- parsing


@dataclass(frozen=True)
class ParsedRow:
    """One data row, already mapped onto canonical field names."""

    row_number: int  # 1-based, counting the header as row 1
    values: dict[str, Any]

    def get(self, key: str) -> Any:
        return self.values.get(key)


@dataclass(frozen=True)
class ParsedFile:
    kind: ExportKind
    rows: list[ParsedRow]
    #: Header text as it appeared, for the audit trail.
    header: list[str]


def _rows_from_matrix(matrix: list[list[Any]]) -> ParsedFile:
    """Shared tail of both parsers: find the header, map every later row."""
    header_idx = None
    for i, row in enumerate(matrix[:20]):  # a report may carry a title row or two
        if any(_norm_header(c) in ("member id", "memberid", "member_id") for c in row):
            header_idx = i
            break
    if header_idx is None:
        raise HeaderValidationError(
            "No header row containing 'Member ID' found in the first 20 rows. "
            "Is this a Yoactiv report export?"
        )

    header = matrix[header_idx]
    kind = detect_kind(header)
    columns = resolve_headers(header, kind)

    rows: list[ParsedRow] = []
    for offset, raw in enumerate(matrix[header_idx + 1 :], start=header_idx + 2):
        if all(str(c or "").strip() == "" for c in raw):
            continue  # trailing blank / spacer row
        values = {name: (raw[i] if i < len(raw) else None) for name, i in columns.items()}
        rows.append(ParsedRow(row_number=offset, values=values))
    return ParsedFile(
        kind=kind, rows=rows, header=[str(c) if c is not None else "" for c in header]
    )


def parse_workbook(path: str | Path, sheet_name: str | None = None) -> ParsedFile:
    """Read a Yoactiv ``.xlsx`` export."""
    from openpyxl import load_workbook

    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
        matrix = [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()
    return _rows_from_matrix(matrix)


def parse_csv(raw: bytes) -> ParsedFile:
    """Read a Yoactiv ``.csv`` export (same reports, CSV flavour)."""
    text = raw.decode("utf-8-sig", errors="replace")
    matrix = [list(r) for r in csv.reader(io.StringIO(text))]
    return _rows_from_matrix(matrix)


class _TableExtractor(HTMLParser):
    """Pull the first real ``<table>`` out of an HTML document as a matrix.

    Yoactiv's "Export Excel" on some reports (the Membership Report among
    them) returns an HTML table served as ``.xls`` — the classic ASP.NET
    ``Response.ContentType = "application/vnd.ms-excel"`` trick. Excel opens
    it, but it is not a workbook and openpyxl cannot read it. Rather than
    demand the operator re-save the file, the bytes are parsed for what they
    actually are.
    """

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.rows: list[list[str]] = []
        self._row: list[str] | None = None
        self._cell: list[str] | None = None

    def handle_starttag(self, tag: str, attrs) -> None:
        if tag == "tr":
            self._row = []
        elif tag in ("td", "th") and self._row is not None:
            self._cell = []
        elif tag == "br" and self._cell is not None:
            self._cell.append(" ")

    def handle_endtag(self, tag: str) -> None:
        if tag in ("td", "th") and self._cell is not None and self._row is not None:
            self._row.append(" ".join("".join(self._cell).split()))
            self._cell = None
        elif tag == "tr" and self._row is not None:
            if any(c for c in self._row):
                self.rows.append(self._row)
            self._row = None

    def handle_data(self, data: str) -> None:
        if self._cell is not None:
            self._cell.append(data)


def parse_html_table(raw: bytes) -> ParsedFile:
    """Read a Yoactiv report exported as an HTML table (served as ``.xls``)."""
    parser = _TableExtractor()
    parser.feed(raw.decode("utf-8", errors="replace"))
    if not parser.rows:
        raise HeaderValidationError(
            "The file looks like HTML but contains no table rows. Re-export the "
            "Yoactiv report, or save it as .xlsx / .csv."
        )
    return _rows_from_matrix([list(r) for r in parser.rows])


def sniff(raw: bytes) -> str:
    """What this file *actually* is, regardless of what it is called.

    Yoactiv names its exports ``.xls`` whether they are a real workbook, an
    HTML table or a CSV, so the extension is not evidence.
    """
    head = raw[:512].lstrip()
    if raw[:2] == b"PK":
        return "xlsx"  # zip container
    if raw[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return "xls"  # OLE2 compound document — a real legacy workbook
    lowered = head[:400].lower()
    if b"<html" in lowered or b"<table" in lowered or lowered.startswith(b"<!doctype"):
        return "html"
    return "csv"


def parse_upload(filename: str, raw: bytes) -> ParsedFile:
    """Parse an uploaded Yoactiv export, whatever container it arrived in."""
    kind = sniff(raw)
    if kind == "xlsx":
        import tempfile

        with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
            tmp.write(raw)
            tmp.flush()
            return parse_workbook(tmp.name)
    if kind == "html":
        return parse_html_table(raw)
    if kind == "csv":
        return parse_csv(raw)
    raise HeaderValidationError(
        f"{filename!r} is a legacy binary .xls workbook, which this importer cannot read. "
        "Open it in Excel or Numbers and save as .xlsx or .csv, then upload again."
    )


# ---------------------------------------------------------------- normalizing


def _clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text in {"-", "--", "NA", "N/A", "None"} else text


def _to_date(value: Any) -> date | None:
    """Yoactiv writes ``dd-MM-yyyy``; openpyxl may hand back a real datetime."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return parse_dmy(value)


def _to_time(value: Any) -> time | None:
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, time):
        return value
    return parse_clock(value)


def _to_amount(value: Any) -> float | None:
    text = _clean(value).replace(",", "").replace("₹", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _member_id(value: Any) -> str:
    """Yoactiv member ids are integers; openpyxl may float them ("2395113.0")."""
    text = _clean(value)
    if not text:
        return ""
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".")[0]
    return text


@dataclass(frozen=True)
class MembershipRecord:
    yoactiv_member_id: str
    name: str
    mobile: str
    email: str
    service_name: str
    starts_on: date
    ends_on: date
    bill_no: str
    bill_amount: float | None
    attendance_id: str
    branch_hint: str


@dataclass(frozen=True)
class CheckinRecord:
    yoactiv_member_id: str
    name: str
    mobile: str
    email: str
    on: date
    clock_in: time | None
    clock_out: time | None
    location_hint: str
    service_name: str
    medium: str

    @property
    def export_key(self) -> str:
        """Dedup identity for an export-sourced visit.

        Deliberately **not** the API's ``mapping.YoactivCheckin.external_key``:
        the export has no ``Service_card_id``, so the two cannot agree. A
        distinct namespace keeps the provenance honest; cross-source
        duplication is prevented by :func:`attendance_natural_key_exists`
        instead of by a key collision that would only look like agreement.
        """
        basis = "|".join(
            (
                self.yoactiv_member_id,
                self.on.isoformat(),
                self.clock_in.isoformat() if self.clock_in else "",
                self.clock_out.isoformat() if self.clock_out else "",
            )
        )
        return "yoactiv:export:checkin:" + hashlib.sha1(basis.encode()).hexdigest()  # noqa: S324


def normalize_membership(row: ParsedRow) -> tuple[MembershipRecord | None, list[str]]:
    errors: list[str] = []
    member_id = _member_id(row.get("member_id"))
    if not member_id:
        errors.append("no Member ID")
    starts_on = _to_date(row.get("start_date"))
    ends_on = _to_date(row.get("end_date"))
    if starts_on is None:
        errors.append("unreadable Start Date")
    if ends_on is None:
        errors.append("unreadable End Date")
    if starts_on and ends_on and ends_on < starts_on:
        errors.append("End Date is before Start Date")
    service = _clean(row.get("service_name"))
    if not service:
        errors.append("no Service Name")
    if errors:
        return None, errors
    assert starts_on is not None and ends_on is not None
    return (
        MembershipRecord(
            yoactiv_member_id=member_id,
            name=_clean(row.get("member_name")),
            mobile=_clean(row.get("mobile")),
            email=_clean(row.get("email")),
            service_name=service,
            starts_on=starts_on,
            ends_on=ends_on,
            bill_no=_clean(row.get("bill_no")),
            bill_amount=_to_amount(row.get("bill_amount")),
            attendance_id=_clean(row.get("attendance_id")),
            branch_hint=_clean(row.get("branch")),
        ),
        [],
    )


def normalize_checkin(row: ParsedRow) -> tuple[CheckinRecord | None, list[str]]:
    errors: list[str] = []
    member_id = _member_id(row.get("member_id"))
    if not member_id:
        errors.append("no Member ID")
    on = _to_date(row.get("date"))
    if on is None:
        errors.append("unreadable Date")
    # Clock times are optional: the on-screen report shows them, the Excel
    # export does not carry them at all. A visit with only a date is still a
    # real visit — see `import_checkins` for how it is filed.
    clock_in = _to_time(row.get("clock_in"))
    clock_out = _to_time(row.get("clock_out"))
    if errors:
        return None, errors
    assert on is not None
    return (
        CheckinRecord(
            yoactiv_member_id=member_id,
            name=_clean(row.get("member_name")),
            mobile=_clean(row.get("mobile")),
            email=_clean(row.get("email")),
            on=on,
            clock_in=clock_in,
            clock_out=clock_out,
            location_hint=_clean(row.get("location")),
            service_name=_clean(row.get("service_name")),
            medium=_clean(row.get("medium")),
        ),
        [],
    )


# -------------------------------------------------------------- classification


class Classification(str, Enum):
    MATCHED = "matched"
    UNMATCHED = "unmatched"
    AMBIGUOUS = "ambiguous"
    DUPLICATE = "duplicate"
    INVALID = "invalid"


@dataclass
class ClassifiedRow:
    row_number: int
    classification: Classification
    detail: str
    member_id: int | None = None
    membership: MembershipRecord | None = None
    checkin: CheckinRecord | None = None
    errors: list[str] = field(default_factory=list)


def attendance_natural_key_exists(
    db: Session, *, user_id: int, work_date: date, event_type: EventType, occurred_at: datetime
) -> bool:
    """Is this exact visit already recorded, whatever wrote it?

    The API connector keys attendance on a composed ``external_event_id`` that
    includes ``Service_card_id``; the export has no such column, so the two
    sources cannot produce the same key for the same visit. Both paths call
    this before inserting so a visit imported from a file is not written again
    by a later API sync (and vice versa).
    """
    return (
        db.scalar(
            select(AttendanceEvent.id).where(
                AttendanceEvent.user_id == user_id,
                AttendanceEvent.work_date == work_date,
                AttendanceEvent.event_type == event_type,
                AttendanceEvent.occurred_at == occurred_at,
            )
        )
        is not None
    )


def _resolve(db: Session, *, member_id: str, name: str, mobile: str, email: str):
    return identity.resolve_member(
        db,
        yoactiv_member_id=member_id,
        email=email or None,
        phone=mobile or None,
        name=name or None,
    )


def classify_membership_rows(db: Session, parsed: ParsedFile) -> list[ClassifiedRow]:
    """Resolve every membership row to a GymFlow member, without writing."""
    out: list[ClassifiedRow] = []
    seen_terms: set[tuple[str, str, date]] = set()
    for row in parsed.rows:
        record, errors = normalize_membership(row)
        if record is None:
            out.append(
                ClassifiedRow(
                    row.row_number, Classification.INVALID, "; ".join(errors), errors=errors
                )
            )
            continue

        term_key = (record.yoactiv_member_id, record.service_name, record.starts_on)
        if term_key in seen_terms:
            out.append(
                ClassifiedRow(
                    row.row_number,
                    Classification.DUPLICATE,
                    "same member/service/start date already seen in this file",
                    membership=record,
                )
            )
            continue
        seen_terms.add(term_key)

        match = _resolve(
            db,
            member_id=record.yoactiv_member_id,
            name=record.name,
            mobile=record.mobile,
            email=record.email,
        )
        if match.method == "ambiguous":
            out.append(
                ClassifiedRow(
                    row.row_number, Classification.AMBIGUOUS, match.detail, membership=record
                )
            )
        elif match.member is None:
            out.append(
                ClassifiedRow(
                    row.row_number, Classification.UNMATCHED, match.detail, membership=record
                )
            )
        else:
            out.append(
                ClassifiedRow(
                    row.row_number,
                    Classification.MATCHED,
                    f"{match.member.member_code} via {match.method}",
                    member_id=match.member.id,
                    membership=record,
                )
            )
    return out


def classify_checkin_rows(db: Session, parsed: ParsedFile) -> list[ClassifiedRow]:
    """Resolve every check-in row, flagging visits already recorded."""
    out: list[ClassifiedRow] = []
    seen_keys: set[str] = set()
    for row in parsed.rows:
        record, errors = normalize_checkin(row)
        if record is None:
            out.append(
                ClassifiedRow(
                    row.row_number, Classification.INVALID, "; ".join(errors), errors=errors
                )
            )
            continue

        if record.export_key in seen_keys:
            out.append(
                ClassifiedRow(
                    row.row_number,
                    Classification.DUPLICATE,
                    "identical visit already seen in this file",
                    checkin=record,
                )
            )
            continue
        seen_keys.add(record.export_key)

        match = _resolve(
            db,
            member_id=record.yoactiv_member_id,
            name=record.name,
            mobile=record.mobile,
            email=record.email,
        )
        if match.method == "ambiguous":
            out.append(
                ClassifiedRow(
                    row.row_number, Classification.AMBIGUOUS, match.detail, checkin=record
                )
            )
        elif match.member is None:
            out.append(
                ClassifiedRow(
                    row.row_number, Classification.UNMATCHED, match.detail, checkin=record
                )
            )
        else:
            out.append(
                ClassifiedRow(
                    row.row_number,
                    Classification.MATCHED,
                    f"{match.member.member_code} via {match.method}",
                    member_id=match.member.id,
                    checkin=record,
                )
            )
    return out


def classify(db: Session, parsed: ParsedFile) -> list[ClassifiedRow]:
    if parsed.kind is ExportKind.MEMBERSHIP:
        return classify_membership_rows(db, parsed)
    return classify_checkin_rows(db, parsed)


def summarize(rows: list[ClassifiedRow]) -> dict[str, int]:
    counts = {c.value: 0 for c in Classification}
    for row in rows:
        counts[row.classification.value] += 1
    counts["total"] = len(rows)
    return counts


# ------------------------------------------------------------------- importing


@dataclass(frozen=True)
class ImportResult:
    kind: ExportKind
    written: int
    skipped: int
    counts: dict[str, int]
    #: Row-level detail for anything a human must look at.
    problems: list[tuple[int, str, str]] = field(default_factory=list)


def _bill_id_for(bill_no: str, member_id: str, starts_on: date) -> int:
    """A stable positive integer for ``YoactivInvoice.bill_id``.

    The export's "Bill No" is a human code ("Aug7-2025"), not the API's
    numeric ``bill_id``. ``lifecycle.apply_invoice`` never reads this field —
    membership rows are keyed on ``(plan_name, starts_on)`` — so a
    deterministic hash keeps the dataclass honest without inventing an
    invoice number that would collide with a real one.
    """
    basis = f"{bill_no}|{member_id}|{starts_on.isoformat()}"
    return int(hashlib.sha1(basis.encode()).hexdigest()[:8], 16)  # noqa: S324


def import_memberships(db: Session, rows: list[ClassifiedRow]) -> ImportResult:
    """Apply MATCHED membership rows through the shared lifecycle.

    Only ``MATCHED`` rows are written. Every membership row goes through
    :func:`lifecycle.apply_invoice`, exactly as an API-synced invoice would,
    so terms upsert on ``(plan_name, starts_on)``, a renewal becomes a new
    row, history is never deleted and ``Member.is_active`` is recomputed the
    same way. Does not commit — the caller owns the transaction.
    """
    written = 0
    problems: list[tuple[int, str, str]] = []
    for row in rows:
        if row.classification is not Classification.MATCHED:
            if row.classification is not Classification.DUPLICATE:
                problems.append((row.row_number, row.classification.value, row.detail))
            continue
        record = row.membership
        assert record is not None and row.member_id is not None
        member = db.get(Member, row.member_id)
        if member is None:  # pragma: no cover - classified moments ago
            problems.append((row.row_number, "invalid", "member disappeared mid-import"))
            continue

        invoice = YoactivInvoice(
            bill_id=_bill_id_for(record.bill_no, record.yoactiv_member_id, record.starts_on),
            member_id=int(record.yoactiv_member_id) if record.yoactiv_member_id.isdigit() else 0,
            name=record.name,
            mobile=record.mobile,
            email=record.email,
            purchase_date=record.starts_on,
            final_amount=record.bill_amount,
            paid_amount=record.bill_amount,
            pt_name="",
            services=(
                YoactivBilledService(
                    description=record.service_name,
                    duration="",
                    base_fee=record.bill_amount,
                    start_date=record.starts_on,
                    end_date=record.ends_on,
                ),
            ),
            raw={"source": SOURCE, "bill_no": record.bill_no},
        )
        change = lifecycle.apply_invoice(db, member, invoice)
        # Stamp the Yoactiv id now that this member is confirmed — later runs
        # then match on external_ref (the strongest key) instead of phone.
        if member.external_ref is None and record.yoactiv_member_id:
            member.external_ref = record.yoactiv_member_id
        if change.changed:
            written += 1
    db.flush()
    return ImportResult(
        kind=ExportKind.MEMBERSHIP,
        written=written,
        skipped=len(rows) - written,
        counts=summarize(rows),
        problems=problems,
    )


def import_checkins(db: Session, rows: list[ClassifiedRow], *, branch: Branch) -> ImportResult:
    """Write MATCHED check-in rows as ``attendance_events``.

    One row can produce two events (a clock-in and a clock-out). Each is
    guarded twice: by its own ``external_event_id`` (this bridge's namespace)
    and by :func:`attendance_natural_key_exists`, which is what stops the API
    connector re-importing the same visit later under a different key.
    Timestamps are filed in the branch's timezone — Yoactiv reports local
    clock times, the same assumption ``sync.py`` documents. Does not commit.
    """
    written = 0
    problems: list[tuple[int, str, str]] = []
    for row in rows:
        if row.classification is not Classification.MATCHED:
            if row.classification is not Classification.DUPLICATE:
                problems.append((row.row_number, row.classification.value, row.detail))
            continue
        record = row.checkin
        assert record is not None and row.member_id is not None
        member = db.get(Member, row.member_id)
        if member is None:  # pragma: no cover
            problems.append((row.row_number, "invalid", "member disappeared mid-import"))
            continue

        # A date-only row (the Excel export carries no clock columns) becomes a
        # single CHECK_IN filed at branch-local midnight, and says so in its
        # notes. Midnight is a marker, not a claim: inventing a plausible
        # 5:42 PM would read as a real recorded time. Every signal that uses
        # attendance — inactivity, consistency, the weekly visit count — keys
        # on `work_date`, which is exact either way.
        dateless = record.clock_in is None and record.clock_out is None
        pairs = (
            (("in", time(0, 0), EventType.CHECK_IN),)
            if dateless
            else (
                ("in", record.clock_in, EventType.CHECK_IN),
                ("out", record.clock_out, EventType.CHECK_OUT),
            )
        )
        for suffix, clock_t, event_type in pairs:
            if clock_t is None:
                continue
            occurred_at = combine_branch(record.on, clock_t, branch.timezone)
            key = f"{record.export_key}:{suffix}"
            already = db.scalar(
                select(AttendanceEvent.id).where(AttendanceEvent.external_event_id == key)
            )
            if already is not None:
                continue
            if attendance_natural_key_exists(
                db,
                user_id=member.user_id,
                work_date=record.on,
                event_type=event_type,
                occurred_at=occurred_at,
            ):
                continue
            db.add(
                AttendanceEvent(
                    branch_id=member.branch_id,
                    person_type=PersonType.MEMBER,
                    user_id=member.user_id,
                    event_type=event_type,
                    method=CaptureMethod.MANUAL,
                    occurred_at=occurred_at,
                    work_date=record.on,
                    device_info=f"{SOURCE}:{record.medium}"[:255] if record.medium else SOURCE,
                    notes=(
                        "Yoactiv export check-in"
                        + (f" ({record.service_name})" if record.service_name else "")
                        + (
                            " — the export carried no clock time; filed at "
                            "branch-local midnight, the visit date is exact"
                            if dateless
                            else ""
                        )
                    )[:500],
                    external_event_id=key,
                )
            )
            db.flush()
            written += 1
    return ImportResult(
        kind=ExportKind.CHECKINS,
        written=written,
        skipped=len(rows) - written,
        counts=summarize(rows),
        problems=problems,
    )


# ------------------------------------------- creating accounts for new members
#
# Opt-in, exactly like the InBody bridge's equivalent: an UNMATCHED row is a
# person Yoactiv knows about and GymFlow does not. Turning one into a GymFlow
# account is a deliberate operator decision, never a side effect of an import.


@dataclass(frozen=True)
class AccountPlan:
    yoactiv_member_id: str
    full_name: str
    phone: str
    row_numbers: list[int]


@dataclass(frozen=True)
class AccountConflict:
    yoactiv_member_id: str
    reason: str  # "shared_phone" | "phone_in_use" | "no_phone" | "name_conflict"
    detail: str
    row_numbers: list[int]


def _record_of(row: ClassifiedRow):
    return row.membership or row.checkin


def plan_accounts(
    db: Session, rows: list[ClassifiedRow]
) -> tuple[list[AccountPlan], list[AccountConflict]]:
    """Which UNMATCHED Yoactiv members could safely become GymFlow accounts.

    Read-only. Eligible: a Yoactiv member id with one name, one usable
    10-digit mobile, and no GymFlow account already on that number. Everything
    else is returned as a conflict for a human — never guessed.
    """
    by_yoactiv_id: dict[str, list[ClassifiedRow]] = defaultdict(list)
    for row in rows:
        if row.classification is Classification.UNMATCHED:
            record = _record_of(row)
            if record is not None:
                by_yoactiv_id[record.yoactiv_member_id].append(row)

    # A mobile number is the Login ID, so two different Yoactiv members on one
    # number cannot both get an account.
    phone_owners: dict[str, set[str]] = defaultdict(set)
    for yid, group in by_yoactiv_id.items():
        for row in group:
            record = _record_of(row)
            if record is not None:
                phone = normalise_phone(record.mobile or "")
                if len(phone) == 10:
                    phone_owners[phone].add(yid)

    plans: list[AccountPlan] = []
    conflicts: list[AccountConflict] = []
    for yid, group in sorted(by_yoactiv_id.items()):
        row_numbers = sorted(r.row_number for r in group)
        records = [_record_of(r) for r in group]
        names = sorted({(r.name or "").strip() for r in records if r and (r.name or "").strip()})
        phones = {normalise_phone(r.mobile or "") for r in records if r}
        phones = {p for p in phones if len(p) == 10}

        if not phones:
            conflicts.append(
                AccountConflict(yid, "no_phone", "no usable 10-digit mobile number", row_numbers)
            )
            continue
        if len(phones) > 1:
            conflicts.append(
                AccountConflict(
                    yid, "shared_phone", "this Yoactiv id has more than one mobile", row_numbers
                )
            )
            continue
        phone = next(iter(phones))
        if len(phone_owners.get(phone, set())) > 1:
            conflicts.append(
                AccountConflict(
                    yid,
                    "shared_phone",
                    "more than one Yoactiv member shares this mobile number",
                    row_numbers,
                )
            )
            continue
        if _user_with_phone(db, phone) is not None:
            conflicts.append(
                AccountConflict(
                    yid,
                    "phone_in_use",
                    "an existing GymFlow account already uses this mobile number",
                    row_numbers,
                )
            )
            continue
        if len(names) > 1:
            conflicts.append(
                AccountConflict(
                    yid,
                    "name_conflict",
                    "different names on one Yoactiv id: " + ", ".join(names),
                    row_numbers,
                )
            )
            continue
        plans.append(
            AccountPlan(
                yoactiv_member_id=yid,
                full_name=names[0] if names else f"Member {phone[-4:]}",
                phone=phone,
                row_numbers=row_numbers,
            )
        )
    return plans, conflicts


def _user_with_phone(db: Session, phone: str) -> User | None:
    hit = db.scalar(select(User).where(User.login_phone == phone))
    if hit is not None:
        return hit
    for candidate in db.scalars(select(User).where(User.phone.isnot(None))).all():
        if normalise_phone(candidate.phone or "") == phone:
            return candidate
    return None


def create_accounts(
    db: Session,
    plans: list[AccountPlan],
    *,
    branch: Branch,
    role_id: int,
    password_hash: str | Callable[[], str],
    joined_on: date,
    email_domain: str = "no-email.gymflow.app",
) -> list[int]:
    """Create a ``User`` + ``Member`` per plan, stamped with the Yoactiv id.

    **No ``Membership`` row** — a membership is created only by importing the
    Membership Report through :func:`import_memberships`, so commercial state
    always traces to a Yoactiv billing row rather than to the mere existence
    of a person.

    The mobile number is the Login ID (``login_phone``). ``password_hash``
    may be a single hash (every account starts on one operator-supplied
    temporary secret) or a **callable invoked per account**, which is how
    the CLI gives each account its own unguessable secret that nobody --
    including this process -- retains. ``must_change_password`` is always
    set. No plaintext enters this module, the database, a log or a
    response. ``is_demo`` stays ``False`` -- these are real people.
    """
    created: list[int] = []
    for plan in plans:
        secret = password_hash() if callable(password_hash) else password_hash
        user = User(
            email=f"{plan.phone}@{email_domain}",
            full_name=plan.full_name[:120],
            phone=plan.phone,
            login_phone=plan.phone,
            password_hash=secret,
            must_change_password=True,
            role_id=role_id,
            branch_id=branch.id,
            is_active=True,
        )
        db.add(user)
        db.flush()
        member = Member(
            user_id=user.id,
            branch_id=branch.id,
            member_code=f"{branch.code}-M{user.id:04d}",
            joined_on=joined_on,
            registered_on=joined_on,
            external_ref=plan.yoactiv_member_id or None,
            is_active=True,
        )
        db.add(member)
        db.flush()
        created.append(member.id)
    db.flush()
    return created


__all__ = [
    "SOURCE",
    "AccountConflict",
    "AccountPlan",
    "CheckinRecord",
    "Classification",
    "ClassifiedRow",
    "ExportKind",
    "HeaderValidationError",
    "ImportResult",
    "MembershipRecord",
    "ParsedFile",
    "ParsedRow",
    "attendance_natural_key_exists",
    "classify",
    "classify_checkin_rows",
    "classify_membership_rows",
    "create_accounts",
    "detect_kind",
    "import_checkins",
    "import_memberships",
    "normalize_checkin",
    "normalize_membership",
    "parse_csv",
    "parse_html_table",
    "parse_upload",
    "parse_workbook",
    "sniff",
    "plan_accounts",
    "resolve_headers",
    "summarize",
]
